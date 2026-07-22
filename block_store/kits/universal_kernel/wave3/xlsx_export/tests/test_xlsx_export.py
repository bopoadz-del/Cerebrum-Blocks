"""Tests for the neutral XLSX export sub-kit."""

import io
import warnings

import pytest
from openpyxl import load_workbook

from block_store.kits.universal_kernel.wave3.xlsx_export import (
    WorkbookBuilder,
    XLSXExportError,
    export_table,
)


def test_workbook_builder_to_bytes():
    builder = WorkbookBuilder()
    builder.add_sheet("data")
    builder.add_header("data", ["id", "name"])
    builder.add_row("data", [1, "alice"])
    raw = builder.to_bytes()
    assert raw.startswith(b"PK")

    wb = load_workbook(io.BytesIO(raw))
    ws = wb["data"]
    assert ws.cell(row=1, column=1).value == "id"
    assert ws.cell(row=2, column=2).value == "alice"


def test_export_table():
    raw = export_table(
        headers=["id", "name"],
        rows=[[1, "alice"], [2, "bob"]],
        title="people",
    )
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["people"]
    assert ws.max_row == 3


def test_empty_export():
    raw = export_table(title="empty")
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["empty"]
    assert ws.max_row == 1


def test_duplicate_sheet_raises():
    builder = WorkbookBuilder()
    builder.add_sheet("data")
    with pytest.raises(XLSXExportError):
        builder.add_sheet("data")


def test_missing_sheet_raises():
    builder = WorkbookBuilder()
    with pytest.raises(XLSXExportError):
        builder.add_row("missing", [1])


def test_invalid_cell_coerced():
    builder = WorkbookBuilder()
    builder.add_sheet("data")
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        builder.add_row("data", [{"nested": True}])
    assert any("complex cell value" in str(w.message) for w in caught)
