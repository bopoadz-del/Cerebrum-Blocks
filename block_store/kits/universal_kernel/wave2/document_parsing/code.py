"""Neutral document parsing primitives: text, CSV, Markdown, JSON, XLSX, PDF."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional


class ParseError(ValueError):
    """Raised when a document cannot be parsed."""


@dataclass
class Document:
    """Neutral parsed document."""

    text: str
    pages: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    chunks: List[str] = field(default_factory=list)
    honesty: str = "parsed"


def _chunk_text(text: str, max_chars: int = 512, overlap: int = 50) -> List[str]:
    """Sliding-window chunker."""
    if not text or not text.strip():
        return []
    if max_chars <= overlap:
        raise ValueError("max_chars must exceed overlap")
    text = text.strip()
    if len(text) <= max_chars:
        return [text]
    chunks: List[str] = []
    step = max_chars - overlap
    start = 0
    while start < len(text):
        end = min(start + max_chars, len(text))
        chunks.append(text[start:end].strip())
        if end >= len(text):
            break
        start += step
    return [c for c in chunks if c]


def _parse_text(content_bytes: bytes, filename: str) -> Document:
    text = content_bytes.decode("utf-8", errors="replace")
    return Document(
        text=text,
        pages=[text],
        metadata={"filename": filename, "type": "text"},
        chunks=_chunk_text(text),
        honesty="parsed",
    )


def _parse_csv(content_bytes: bytes, filename: str) -> Document:
    text = content_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    lines = [", ".join(row) for row in rows]
    parsed_text = "\n".join(lines)
    return Document(
        text=parsed_text,
        pages=[parsed_text],
        metadata={"filename": filename, "type": "csv", "row_count": max(0, len(rows) - 1) if rows else 0},
        chunks=_chunk_text(parsed_text),
        honesty="parsed",
    )


def _parse_json(content_bytes: bytes, filename: str) -> Document:
    text = content_bytes.decode("utf-8", errors="replace")
    data = json.loads(text)
    pretty = json.dumps(data, ensure_ascii=False, indent=2)
    return Document(
        text=pretty,
        pages=[pretty],
        metadata={"filename": filename, "type": "json"},
        chunks=_chunk_text(pretty),
        honesty="parsed",
    )


def _parse_xlsx(content_bytes: bytes, filename: str) -> Document:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    all_lines: List[str] = []
    total_rows = 0
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell) if cell is not None else "" for cell in rows[0]]
        for row in rows[1:]:
            total_rows += 1
            cells = [str(cell) if cell is not None else "" for cell in row]
            pairs = [f"{header[i]}={cells[i]}" for i in range(min(len(header), len(cells)))]
            if any(cells):
                all_lines.append(f"[{sheet.title}] " + ", ".join(pairs))
    parsed_text = "\n".join(all_lines)
    return Document(
        text=parsed_text,
        pages=[parsed_text],
        metadata={"filename": filename, "type": "xlsx", "row_count": total_rows},
        chunks=_chunk_text(parsed_text),
        honesty="parsed",
    )


def _parse_pdf(content_bytes: bytes, filename: str) -> Document:
    try:
        from pypdf import PdfReader
    except ImportError:  # pragma: no cover - optional dependency
        raise ParseError("pypdf is not installed")

    reader = PdfReader(io.BytesIO(content_bytes))
    pages: List[str] = []
    for page in reader.pages:
        page_text = (page.extract_text() or "").strip()
        if page_text:
            pages.append(page_text)
    text = "\n\n".join(pages)
    return Document(
        text=text,
        pages=pages,
        metadata={"filename": filename, "type": "pdf", "page_count": len(pages)},
        chunks=_chunk_text(text),
        honesty="parsed",
    )


_PARSER_MAP = {
    "text/plain": _parse_text,
    "text/csv": _parse_csv,
    "text/markdown": _parse_text,
    "application/json": _parse_json,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _parse_xlsx,
    "application/pdf": _parse_pdf,
}


def parse(
    content_bytes: bytes,
    mime_type: str,
    filename: str,
    extract_raw: bool = False,
) -> Document:
    """Parse content into a neutral Document; fail closed on unsupported types."""
    if content_bytes is None:
        raise ParseError("content_bytes is required")

    parser = _PARSER_MAP.get(mime_type)
    if parser is None:
        if extract_raw:
            return Document(
                text="",
                pages=[],
                metadata={"filename": filename, "type": mime_type, "error": "unsupported mime type"},
                chunks=[],
                honesty="fallback",
            )
        raise ParseError(f"unsupported mime type: {mime_type}")

    try:
        return parser(content_bytes, filename)
    except Exception as exc:
        if extract_raw:
            return Document(
                text="",
                pages=[],
                metadata={"filename": filename, "type": mime_type, "error": str(exc)},
                chunks=[],
                honesty="fallback",
            )
        raise
