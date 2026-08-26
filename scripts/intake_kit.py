#!/usr/bin/env python3
"""Stage K1 of the kit pipeline: an encoding sheet in, a sourced intake out.

WHAT THIS REFUSES, AND WHY THAT IS THE POINT
--------------------------------------------
"No source, no encode." An item that does not say where it came from is
dropped and listed in the report. It is never encoded with a blank source,
because a kit full of unattributed assertions is indistinguishable from a kit
full of invented ones once it is installed -- and the reader of a generated
product cannot tell which they have.

This is the same rule ``audit_kit_composition`` applies to composition and
``formula_definitions`` applies to arithmetic: an absent declaration is not
an empty one, and silence is not permission.

WHAT THIS DOES NOT GUESS
------------------------
There is no encoding sheet in either repository, and no committed schema for
one. So this tool does not invent section semantics:

* Sections are matched by heading text against :data:`SECTION_ALIASES`, which
  is a starting map, not a specification. Override it with ``--section-map``.
* A section it cannot map is **reported as unmapped**, not filed under a
  best guess. Filing an unrecognised section under "vocabulary" because the
  word looked similar would put content in a kit that nobody chose to put
  there.
* Metadata columns are matched by header name. A table whose headers it does
  not recognise is reported, not parsed positionally -- column order is not
  a contract anyone wrote down.
* ``confidence`` is carried through when the sheet states it and left
  ``null`` when it does not. It is never synthesised: a made-up 0.8 is worse
  than an honest blank, because it survives into the kit looking measured.

The one section number this file hardcodes -- Section 17 for refusal traps --
comes from the pipeline spec, not from inspection of any document.

Usage:
    python scripts/intake_kit.py --sheet dental.xlsx --kit dental \\
        --contributor dr-a-hassan
    python scripts/intake_kit.py --sheet dental.docx --kit dental \\
        --contributor dr-a-hassan --out intake/dental.yaml
    python scripts/intake_kit.py --sheet already_structured.yaml --kit dental \\
        --contributor dr-a-hassan --strict

Exit codes:
    0  intake written; every item carried a source
    1  intake written, but items were dropped (or --strict and anything was)
    2  usage error, unreadable sheet, or nothing recognisable in it
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]

#: The categories a kit is built from. An item lands in exactly one.
SECTIONS = (
    "entities",
    "workflows",
    "formulas",
    "precedence",
    "vocabulary",
    "refusal_traps",
)

#: Heading text -> section. A STARTING MAP, not a schema: no encoding sheet
#: exists to derive one from. Anything unmatched is reported as unmapped
#: rather than filed under a guess. Override wholesale with --section-map.
SECTION_ALIASES: Dict[str, str] = {
    "entities": "entities",
    "entity": "entities",
    "objects": "entities",
    "data model": "entities",
    "workflows": "workflows",
    "workflow": "workflows",
    "processes": "workflows",
    "procedures": "workflows",
    "formulas": "formulas",
    "formula": "formulas",
    "calculations": "formulas",
    "metrics": "formulas",
    "precedence": "precedence",
    "precedence rules": "precedence",
    "authority": "precedence",
    "conflict resolution": "precedence",
    "vocabulary": "vocabulary",
    "vocab": "vocabulary",
    "glossary": "vocabulary",
    "terminology": "vocabulary",
    "refusal traps": "refusal_traps",
    "refusals": "refusal_traps",
    "failure modes": "refusal_traps",
    # Section 17 is named by the pipeline spec as the refusal-trap section.
    # It is recorded here because the spec says so, not because any document
    # was inspected.
    "section 17": "refusal_traps",
}

#: Provenance kinds an item may claim. Anything else is refused rather than
#: coerced -- "internal doc" is not a synonym for "standard".
SOURCE_KINDS = ("regulator", "standard", "company_policy", "contributor")

#: Column/field headers recognised as metadata rather than content.
_FIELD_ALIASES = {
    "source": "source",
    "source kind": "source",
    "source_type": "source",
    "source ref": "source_ref",
    "source_ref": "source_ref",
    "reference": "source_ref",
    "citation": "source_ref",
    "confidence": "confidence",
    "item": "text",
    "text": "text",
    "statement": "text",
    "description": "text",
    "rule": "text",
}


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _norm_key(value: Any) -> str:
    return _norm(value).lower().rstrip(":").strip()


class IntakeItem(dict):
    """One extracted assertion plus where it came from."""

    @property
    def has_source(self) -> bool:
        return bool(self.get("source")) and bool(self.get("source_ref"))


def _make_item(
    section: str,
    text: str,
    location: str,
    fields: Dict[str, Any],
    contributor_id: str,
) -> IntakeItem:
    source = _norm_key(fields.get("source"))
    confidence = fields.get("confidence")
    if confidence is not None and _norm(confidence) == "":
        confidence = None
    return IntakeItem(
        section=section,
        text=text,
        location=location,
        source=source if source in SOURCE_KINDS else (source or ""),
        source_ref=_norm(fields.get("source_ref")),
        contributor_id=contributor_id,
        # Never synthesised. A blank stays blank.
        confidence=confidence,
    )


# -- readers ---------------------------------------------------------------


def _map_section(heading: str, aliases: Dict[str, str]) -> Optional[str]:
    key = _norm_key(heading)
    if key in aliases:
        return aliases[key]
    # "3. Formulas" / "Section 4 - Workflows": strip a leading ordinal.
    stripped = re.sub(r"^(section\s*)?\d+\s*[.\-:)]?\s*", "", key).strip()
    return aliases.get(stripped)


def _header_map(row: Iterable[Any]) -> Optional[Dict[int, str]]:
    """Column index -> canonical field, or None if the header is unrecognised.

    Returning None is deliberate. Parsing an unknown header positionally
    would assign somebody's ``notes`` column to ``source_ref`` and produce a
    citation that was never written.
    """
    mapping: Dict[int, str] = {}
    for index, cell in enumerate(row):
        field = _FIELD_ALIASES.get(_norm_key(cell))
        if field:
            mapping[index] = field
    if "text" not in mapping.values():
        return None
    return mapping


def _rows_to_items(
    rows: List[List[Any]],
    section: str,
    where: str,
    contributor_id: str,
) -> Tuple[List[IntakeItem], Optional[str]]:
    if not rows:
        return [], f"{where}: empty"
    header = _header_map(rows[0])
    if header is None:
        return [], (
            f"{where}: header not recognised "
            f"({', '.join(_norm(c) for c in rows[0] if _norm(c))[:80]})"
        )
    items: List[IntakeItem] = []
    for offset, row in enumerate(rows[1:], start=2):
        fields = {
            header[i]: row[i]
            for i in header
            if i < len(row) and _norm(row[i]) != ""
        }
        text = _norm(fields.pop("text", ""))
        if not text:
            continue
        items.append(
            _make_item(section, text, f"{where} row {offset}", fields, contributor_id)
        )
    return items, None


def read_xlsx(path: Path, aliases: Dict[str, str], contributor_id: str):
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=True)
    items: List[IntakeItem] = []
    unmapped: List[str] = []
    for sheet in book.worksheets:
        section = _map_section(sheet.title, aliases)
        if section is None:
            unmapped.append(f"sheet {sheet.title!r}: no section mapping")
            continue
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        found, note = _rows_to_items(
            rows, section, f"{path.name}!{sheet.title}", contributor_id
        )
        items.extend(found)
        if note:
            unmapped.append(note)
    return items, unmapped


def read_docx(path: Path, aliases: Dict[str, str], contributor_id: str):
    import docx  # python-docx

    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = docx.Document(str(path))
    items: List[IntakeItem] = []
    unmapped: List[str] = []

    # Walk the body in document order so a table is attributed to the heading
    # it actually sits under. Pairing table N with heading N only holds when
    # every heading owns exactly one table, which no real document promises.
    #
    # Prose is deliberately not mined for assertions: a paragraph has no
    # source column, so anything harvested from it would be dropped by the
    # gate anyway -- and silently mining it would invite someone to "fix" the
    # gate rather than write the source down.
    heading = ""
    section: Optional[str] = None
    seen_headings: set = set()
    table_index = 0

    for child in document.element.body.iterchildren():
        tag = child.tag.split("}")[-1]
        if tag == "p":
            block = Paragraph(child, document)
            style = (block.style.name if block.style else "") or ""
            if style.lower().startswith("heading") and _norm(block.text):
                heading = _norm(block.text)
                section = _map_section(heading, aliases)
                if section is None and heading not in seen_headings:
                    seen_headings.add(heading)
                    unmapped.append(f"heading {heading!r}: no section mapping")
        elif tag == "tbl":
            table_index += 1
            table = Table(child, document)
            if section is None:
                unmapped.append(
                    f"table {table_index} under {heading or '<no heading>'!r}: "
                    "no section mapping"
                )
                continue
            rows = [[cell.text for cell in row.cells] for row in table.rows]
            found, note = _rows_to_items(
                rows, section, f"{path.name} table {table_index}", contributor_id
            )
            items.extend(found)
            if note:
                unmapped.append(note)
    return items, unmapped


def read_structured(path: Path, aliases: Dict[str, str], contributor_id: str):
    """A sheet somebody already converted: {section: [ {...}, ... ]}."""
    text = path.read_text(encoding="utf-8")
    data = json.loads(text) if path.suffix.lower() == ".json" else _yaml_load(text)
    if not isinstance(data, dict):
        return [], [f"{path.name}: top level is not a mapping of sections"]

    items: List[IntakeItem] = []
    unmapped: List[str] = []
    for heading, entries in data.items():
        section = _map_section(heading, aliases)
        if section is None:
            unmapped.append(f"key {heading!r}: no section mapping")
            continue
        if not isinstance(entries, list):
            unmapped.append(f"key {heading!r}: expected a list of items")
            continue
        for position, entry in enumerate(entries, start=1):
            if not isinstance(entry, dict):
                unmapped.append(f"{heading}[{position}]: expected a mapping")
                continue
            fields = {
                _FIELD_ALIASES.get(_norm_key(k), _norm_key(k)): v
                for k, v in entry.items()
            }
            text_value = _norm(fields.pop("text", ""))
            if not text_value:
                unmapped.append(f"{heading}[{position}]: no item text")
                continue
            items.append(
                _make_item(
                    section,
                    text_value,
                    f"{path.name}:{heading}[{position}]",
                    fields,
                    contributor_id,
                )
            )
    return items, unmapped


def _yaml_load(text: str):
    import yaml

    return yaml.safe_load(text)


_READERS = {
    ".xlsx": read_xlsx,
    ".xlsm": read_xlsx,
    ".docx": read_docx,
    ".yaml": read_structured,
    ".yml": read_structured,
    ".json": read_structured,
}


# -- the gate --------------------------------------------------------------


def apply_source_gate(items: List[IntakeItem]):
    """Split into (encoded, dropped). No source, no encode.

    An item is dropped when it names no ``source``, names one outside
    :data:`SOURCE_KINDS`, or gives no ``source_ref``. Every drop is returned
    with the reason and its location so the contributor can fix the sheet
    rather than guess what the tool disliked.
    """
    kept: List[IntakeItem] = []
    dropped: List[Dict[str, str]] = []
    for item in items:
        if not item.get("source"):
            reason = "no source named"
        elif item["source"] not in SOURCE_KINDS:
            reason = (
                f"source {item['source']!r} is not one of "
                f"{', '.join(SOURCE_KINDS)}"
            )
        elif not item.get("source_ref"):
            reason = "source named but no reference given"
        else:
            kept.append(item)
            continue
        dropped.append(
            {
                "location": item["location"],
                "section": item["section"],
                "text": item["text"][:120],
                "reason": reason,
            }
        )
    return kept, dropped


def build_intake(
    kit: str,
    sheet: Path,
    items: List[IntakeItem],
    dropped: List[Dict[str, str]],
    unmapped: List[str],
    contributor_id: str,
    ingested_at: str,
) -> Dict[str, Any]:
    by_section: Dict[str, List[Dict[str, Any]]] = {s: [] for s in SECTIONS}
    for item in items:
        by_section[item["section"]].append(dict(item))

    unsourced_confidence = sum(1 for i in items if i.get("confidence") is None)
    return {
        "schema_version": 1,
        "kit": kit,
        "intake": {
            "sheet": sheet.name,
            "sheet_sha256": hashlib.sha256(sheet.read_bytes()).hexdigest(),
            "contributor_id": contributor_id,
            "ingested_at": ingested_at,
            "tool": "scripts/intake_kit.py",
        },
        "sections": by_section,
        "report": {
            "encoded": len(items),
            "dropped": dropped,
            "unmapped": unmapped,
            "without_confidence": unsourced_confidence,
            "note": (
                "Dropped items were not encoded. They are listed so the sheet "
                "can be corrected -- no source, no encode. Unmapped sections "
                "were left out rather than filed under a guess."
            ),
        },
    }


def write_intake(document: Dict[str, Any], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.suffix.lower() == ".json":
        out.write_text(json.dumps(document, indent=2) + "\n", encoding="utf-8")
        return
    import yaml

    out.write_text(
        yaml.safe_dump(document, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Parse an encoding sheet into a sourced kit intake (stage K1)."
    )
    parser.add_argument("--sheet", required=True, type=Path)
    parser.add_argument("--kit", required=True, help="Kit id the intake is for")
    parser.add_argument(
        "--contributor",
        required=True,
        help="Contributor identity recorded on every extracted item",
    )
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument(
        "--section-map",
        type=Path,
        default=None,
        help="JSON mapping of heading text -> section, replacing the default",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit non-zero if anything was dropped or left unmapped",
    )
    parser.add_argument(
        "--ingested-at",
        default=None,
        help="ISO timestamp to record (defaults to now, UTC)",
    )
    args = parser.parse_args()

    sheet: Path = args.sheet
    if not sheet.is_file():
        print(f"No such sheet: {sheet}", file=sys.stderr)
        return 2

    reader = _READERS.get(sheet.suffix.lower())
    if reader is None:
        print(
            f"Cannot read {sheet.suffix!r}. Supported: "
            f"{', '.join(sorted(_READERS))}",
            file=sys.stderr,
        )
        return 2

    aliases = dict(SECTION_ALIASES)
    if args.section_map:
        override = json.loads(args.section_map.read_text(encoding="utf-8"))
        bad = sorted({v for v in override.values() if v not in SECTIONS})
        if bad:
            print(
                f"--section-map targets sections that do not exist: {bad}. "
                f"Known: {', '.join(SECTIONS)}",
                file=sys.stderr,
            )
            return 2
        aliases = {_norm_key(k): v for k, v in override.items()}

    try:
        items, unmapped = reader(sheet, aliases, args.contributor)
    except ImportError as exc:
        print(f"Reader for {sheet.suffix} needs a package that is not installed: {exc}",
              file=sys.stderr)
        return 2

    if not items and not unmapped:
        print(
            f"Nothing recognisable in {sheet.name}. Either it has no mapped "
            f"sections or no tables with a recognised header.",
            file=sys.stderr,
        )
        return 2

    kept, dropped = apply_source_gate(items)
    ingested_at = args.ingested_at or datetime.now(timezone.utc).isoformat()
    document = build_intake(
        args.kit, sheet, kept, dropped, unmapped, args.contributor, ingested_at
    )

    out = args.out or (PROJECT_ROOT / "intake" / f"{args.kit}.yaml")
    write_intake(document, out)

    print(f"Intake written: {out}")
    print(f"  encoded: {len(kept)}")
    for section in SECTIONS:
        count = len(document["sections"][section])
        if count:
            print(f"    {section}: {count}")
    if dropped:
        print(f"  DROPPED (no source, no encode): {len(dropped)}")
        for entry in dropped[:10]:
            print(f"    {entry['location']}: {entry['reason']}")
        if len(dropped) > 10:
            print(f"    ... and {len(dropped) - 10} more, all listed in the intake")
    if unmapped:
        print(f"  unmapped: {len(unmapped)}")
        for note in unmapped[:10]:
            print(f"    {note}")
    if document["report"]["without_confidence"]:
        print(
            f"  without a stated confidence: "
            f"{document['report']['without_confidence']} "
            f"(left null rather than invented)"
        )

    if dropped or (args.strict and unmapped):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
