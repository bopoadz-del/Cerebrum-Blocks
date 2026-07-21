"""PDF Block - Extract text from PDF files with Typed Schema"""

import logging
import os
import tempfile
from typing import Any, Dict, Optional
from app.core.typed_block import TypedBlock, Schema, ContentType

logger = logging.getLogger(__name__)


def _try_marker(pdf_path: str) -> Optional[Dict[str, Any]]:
    """Convert PDF to Markdown using Marker if available.

    Returns a dict with ``text`` and ``pages`` on success, or ``None`` if
    marker-pdf is not installed or fails. This lets the existing PDF block
    opt into high-quality extraction via ``params={"use_marker": true}``.
    """
    try:
        from marker.converters.pdf import PdfConverter
        from marker.models import create_model_dict
        from marker.output import text_from_rendered
    except ImportError as e:
        logger.info("Marker not installed: %s", e)
        return None

    try:
        converter = PdfConverter(artifact_dict=create_model_dict())
        rendered = converter(pdf_path)
        text, _, _ = text_from_rendered(rendered)
        pages = getattr(rendered, "page_count", len(getattr(rendered, "pages", [])))
        return {"text": text or "", "pages": pages or 1}
    except Exception as e:
        logger.warning("Marker conversion failed for %s: %s", pdf_path, e)
        return None


def _try_ocr_fallback(
    pdf_path: str,
    original_text: str,
    max_pages: int = 30,
    dpi: int = 200,
    time_budget_s: float = 60.0,
) -> Optional[str]:
    """Rasterise a scanned PDF and OCR each page until we beat the original
    text-extraction result, hit ``max_pages``, or burn ``time_budget_s``.

    Returns the OCR'd text only if it has more non-whitespace content than
    ``original_text``; otherwise ``None`` (signalling to the caller to keep
    the original result). Returns ``None`` on any import/runtime failure —
    OCR is best-effort, never required.
    """
    try:
        import io
        import time

        import fitz  # PyMuPDF
        import pytesseract
        from PIL import Image
    except ImportError as e:  # pragma: no cover - exercised via monkeypatch
        logger.warning("OCR fallback skipped: %s", e)
        return None

    started = time.perf_counter()
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logger.warning("OCR fallback could not open %s: %s", pdf_path, e)
        return None

    try:
        ocr_pages = []
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            if time.perf_counter() - started > time_budget_s:
                logger.warning("OCR fallback hit time budget at page %d", i)
                break
            try:
                pix = page.get_pixmap(dpi=dpi)
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                page_text = pytesseract.image_to_string(img, timeout=10) or ""
            except Exception as e:
                logger.warning("OCR failed on page %d: %s", i, e)
                continue
            if page_text.strip():
                ocr_pages.append(f"\n--- page {i + 1} ---\n{page_text}")
        out = "".join(ocr_pages)
        if len(out.strip()) > len(original_text.strip()):
            return out
        return None
    finally:
        doc.close()


class PDFBlock(TypedBlock):
    """Extract text from PDF files with typed output"""
    
    name = "pdf"
    version = "2.0.0"
    description = "Extract text from PDF files"
    layer = 3
    tags = ["domain", "documents", "pdf", "typed"]
    requires = []
    
    default_config = {
        "extract_tables": True
    }
    
    # Type schemas for chain validation
    input_schema = Schema(
        content_type=ContentType.FILE,
        required_fields=["file_path"],
        optional_fields=["path", "url"],
        format_hints={"accept": [".pdf"]}
    )
    
    output_schema = Schema(
        content_type=ContentType.PDF,
        required_fields=["text"],
        optional_fields=["pages", "filename", "status"],
        format_hints={"max_chars": 20000}
    )
    
    ui_schema = {
        "input": {
            "type": "file",
            "accept": [".pdf"],
            "placeholder": "Upload PDF...",
            "multiline": False
        },
        "output": {
            "type": "json",
            "fields": [
                {"name": "text", "type": "text", "label": "Text"},
                {"name": "pages", "type": "number", "label": "Pages"}
            ]
        },
        "quick_actions": [
            {"icon": "📄", "label": "Extract Text", "prompt": "Extract all text from this PDF"},
            {"icon": "📊", "label": "Extract Tables", "prompt": "Extract all tables from this PDF as structured data"},
            {"icon": "📝", "label": "Summarize", "prompt": "Summarize the key points of this document"}
        ]
    }
    
    async def process(self, input_data: Any, params: Dict = None) -> Dict:
        """Extract text from PDF"""
        params = params or {}
        
        # If input is a URL, download it first
        url = None
        if isinstance(input_data, dict):
            url = input_data.get("url")
            # InputAdapter wraps bare strings as {"text": "..."} — check for URL there
            if not url:
                raw = input_data.get("text") or input_data.get("input") or ""
                if raw.startswith("http"):
                    url = raw
        elif isinstance(input_data, str) and input_data.startswith("http"):
            url = input_data
        
        if url:
            import httpx
            try:
                async with httpx.AsyncClient(follow_redirects=True) as client:
                    response = await client.get(url, timeout=30)
                    response.raise_for_status()
                    suffix = ".pdf" if ".pdf" in url.lower() else ".tmp"
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                        f.write(response.content)
                        input_data = f.name
            except Exception as e:
                return {"status": "error", "text": "", "pages": 0, "error": f"Download failed: {str(e)}"}
        
        # Get PDF path (handles bytes, strings, dicts)
        pdf_path = self._get_pdf_path(input_data)
        if not pdf_path:
            return {"status": "error", "text": "", "pages": 0, "error": "No PDF provided"}

        if not os.path.exists(pdf_path):
            return {"status": "error", "text": "", "pages": 0, "error": f"File not found: {pdf_path}"}

        ext = os.path.splitext(pdf_path)[1].lower()

        # Optional Marker pass: high-quality PDF -> Markdown.
        if params.get("use_marker") and ext == ".pdf":
            marker_result = _try_marker(pdf_path)
            if marker_result is not None:
                return {
                    "status": "success",
                    "text": marker_result["text"][:200000],
                    "pages": marker_result["pages"],
                    "filename": os.path.basename(pdf_path),
                    "file_path": pdf_path,
                    "engine": "marker",
                }

        # Handle Excel spreadsheets (XLS/XLSX)
        if ext in ('.xls', '.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(pdf_path, read_only=True, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    parts.append(f"=== Sheet: {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True):
                        row_text = '\t'.join(str(c) if c is not None else '' for c in row)
                        if row_text.strip():
                            parts.append(row_text)
                wb.close()
                text = '\n'.join(parts)
                return {
                    "status": "success",
                    "text": text[:200000],
                    "pages": len(wb.sheetnames) if hasattr(wb, 'sheetnames') else 1,
                    "filename": os.path.basename(pdf_path),
                    "file_path": pdf_path,
                    "engine": "openpyxl",
                }
            except Exception as e:
                return {"status": "error", "text": "", "pages": 0, "error": f"Excel read failed: {str(e)}"}

        # Handle Word documents (DOCX)
        if ext in ('.doc', '.docx'):
            try:
                import docx as python_docx
                doc = python_docx.Document(pdf_path)
                text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
                return {
                    "status": "success",
                    "text": text[:200000],
                    "pages": 1,
                    "filename": os.path.basename(pdf_path),
                    "file_path": pdf_path,
                    "engine": "python-docx",
                }
            except ImportError:
                pass  # Fall through — python-docx not installed
            except Exception as e:
                return {"status": "error", "text": "", "pages": 0, "error": f"Word read failed: {str(e)}"}

        # Try real PDF libraries in order: pdfplumber, pypdf, PyMuPDF.
        # On success, fall through to the OCR-fallback gate below instead
        # of returning early — a vector/scanned PDF will extract cleanly
        # but produce almost no characters per page, and we want to retry
        # with tesseract before we hand the user a near-empty result.
        last_error = None
        text = None
        pages = 0
        engine = None

        # 1. Try pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                buf = ""
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    buf += page_text + "\n"
                text = buf
                pages = len(pdf.pages)
                engine = "pdfplumber"
        except ImportError:
            last_error = "pdfplumber not installed"
        except Exception as e:
            last_error = f"pdfplumber failed: {str(e)}"

        # 2. Try pypdf
        if text is None:
            try:
                from pypdf import PdfReader
                reader = PdfReader(pdf_path)
                buf = ""
                for page in reader.pages:
                    page_text = page.extract_text() or ""
                    buf += page_text + "\n"
                text = buf
                pages = len(reader.pages)
                engine = "pypdf"
            except ImportError:
                last_error = "pypdf not installed"
            except Exception as e:
                last_error = f"pypdf failed: {str(e)}"

        # 3. Try PyMuPDF fallback
        if text is None:
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(pdf_path)
                buf = ""
                for page in doc:
                    buf += page.get_text()
                text = buf
                pages = len(doc)
                doc.close()
                engine = "PyMuPDF"
            except ImportError:
                last_error = "PyMuPDF not installed"
            except Exception as e:
                last_error = f"PyMuPDF failed: {str(e)}"

        if text is None:
            return {
                "status": "error",
                "text": "",
                "pages": 0,
                "error": (
                    f"No PDF parser available. {last_error}. "
                    "Install one of: pdfplumber, pypdf, or PyMuPDF"
                ),
            }

        # OCR fallback gate: only fires for actual .pdf files when the
        # text-density-per-page is suspiciously low. The 100-char/page
        # threshold is the minimum where a real document stops looking
        # like a scan or a vector-only drawing.
        if ext == ".pdf":
            text_density = len(text.strip()) / max(pages, 1)
            if text_density < 100:
                ocr_text = _try_ocr_fallback(pdf_path, text)
                if ocr_text is not None:
                    if len(text.strip()) == 0:
                        engine = "OCR-only"
                    else:
                        engine = f"{engine} + OCR"
                    text = ocr_text

        return {
            "status": "success",
            "text": text[:200000],
            "pages": pages,
            "filename": os.path.basename(pdf_path),
            "file_path": pdf_path,
            "engine": engine,
        }
    
    def _get_pdf_path(self, input_data: Any) -> str:
        """Extract PDF path from input, writing bytes to a temp file if needed."""
        if isinstance(input_data, bytes):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                f.write(input_data)
                return f.name
        if isinstance(input_data, str):
            return input_data
        if isinstance(input_data, dict):
            # Check for explicit file bytes inside dict
            file_bytes = input_data.get("file") or input_data.get("pdf_bytes") or input_data.get("bytes")
            if isinstance(file_bytes, bytes):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                    f.write(file_bytes)
                    return f.name
            return input_data.get("file_path") or input_data.get("path") or input_data.get("url")
        return None
