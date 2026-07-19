"""Tests for app.lib.pm_excel."""

import openpyxl

from app.lib.pm_excel import generate_cost_loaded_schedule, generate_evm_workbook


def _eval_cell(ws, coord):
    v = ws[coord].value
    if isinstance(v, str) and v.startswith("="):
        # Minimal evaluator for the formulas we emit.
        expr = v[1:]
        if expr.startswith("SUM("):
            rng = expr[4:-1]
            start, end = rng.split(":")
            return sum(_eval_cell(ws, c) for c in _range_cells(start, end))
        if "!" in expr:
            return 0.0  # cross-sheet not needed for these assertions
        if "*" in expr:
            a, b = expr.split("*")
            return _eval_cell(ws, a) * _eval_cell(ws, b)
        if "+" in expr:
            a, b = expr.split("+")
            return _eval_cell(ws, a) + _eval_cell(ws, b)
        if "-" in expr:
            a, b = expr.split("-")
            return _eval_cell(ws, a) - _eval_cell(ws, b)
        if "/" in expr:
            a, b = expr.split("/")
            return _eval_cell(ws, a) / _eval_cell(ws, b)
        return float(ws[expr].value or 0)
    return float(v or 0)


def _range_cells(start, end):
    from openpyxl.utils import get_column_letter

    c1 = ord(start[0]) - ord("A") + 1
    c2 = ord(end[0]) - ord("A") + 1
    r1 = int(start[1:])
    r2 = int(end[1:])
    return [f"{get_column_letter(c)}{r}" for c in range(c1, c2 + 1) for r in range(r1, r2 + 1)]


def test_generate_cost_loaded_schedule_default_currency_is_usd():
    meta = {"project": "Demo"}
    activities = [
        {"id": "1", "name": "Start", "duration": 0, "manpower": 0, "predecessors": []},
    ]
    wb = generate_cost_loaded_schedule(meta, activities)
    ws = wb["L2 Schedule"]
    assert "Cost (USD)" in [ws.cell(row=3, column=c).value for c in range(1, 13)]


def test_generate_cost_loaded_schedule_with_activities():
    meta = {"project": "Demo", "currency": "USD", "start_date": "2026-01-01"}
    activities = [
        {"id": "A", "name": "Excavation", "duration": 5, "manpower": 4, "wbs": "E", "cost": 1000, "predecessors": []},
        {"id": "B", "name": "Concrete", "duration": 7, "manpower": 6, "wbs": "C", "cost": 2000, "predecessors": ["A"]},
    ]
    wb = generate_cost_loaded_schedule(meta, activities)
    assert "L2 Schedule" in wb.sheetnames
    assert "Cost Loading" in wb.sheetnames
    assert "Manpower Histogram" in wb.sheetnames
    assert "Milestones" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    sched = wb["L2 Schedule"]
    assert sched["A1"].value == "L2 SCHEDULE — Demo"
    # Row 4 = first activity A, row 5 = B
    assert sched["A4"].value == "A"
    assert sched["D4"].value == 5
    assert sched["L4"].value == 1000
    assert sched["L5"].value == 2000
    total = _eval_cell(sched, "L6")
    assert total == 3000


def test_generate_evm_workbook():
    meta = {"project": "Demo", "bac": 100000, "currency": "USD"}
    periods = [
        {"period": 1, "pv": 20000, "ev": 18000, "ac": 19000},
        {"period": 2, "pv": 40000, "ev": 42000, "ac": 41000},
    ]
    wb = generate_evm_workbook(meta, periods)
    ws = wb["EVM"]
    assert ws["A1"].value == "EARNED VALUE MANAGEMENT — Demo"
    assert ws["B2"].value == 100000
    # Row 5 = period 1
    assert ws["A5"].value == 1
    assert ws["B5"].value == 20000
    assert ws["E5"].value == "=C5-D5"
    cv = _eval_cell(ws, "E5")
    assert cv == -1000
    sv = _eval_cell(ws, "F5")
    assert sv == -2000
