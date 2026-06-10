"""Regression tests for XLSX schedule extraction.

Real-world driver: owners commonly issue L2/L3 schedules as Excel rather
than Primavera XER/XML. These tests exercise the heuristic XLSX parser
that lives on `ConstructionContainer._parse_xlsx_schedule` and confirm
the dispatch wiring in `parse_primavera_schedule` and `_classify_document`.
"""

from __future__ import annotations

import asyncio
import os
import tempfile
from datetime import datetime

import pytest
from openpyxl import Workbook

from tests.conftest import CONSTRUCTION_CONTAINER_PATH

if not CONSTRUCTION_CONTAINER_PATH.exists():
    pytest.skip(
        "Construction kit not installed — run store install or copy from "
        "block_store/kits/construction/bundle/",
        allow_module_level=True,
    )

from app.containers.construction import ConstructionContainer


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _write_workbook(wb: Workbook) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def _basic_schedule_workbook() -> Workbook:
    """1 sheet titled 'Project Schedule' with 5 activity rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Schedule"
    ws.append(["Activity ID", "Activity Name", "Duration", "Start", "Finish", "% Complete"])
    rows = [
        ("A1000", "Site Mobilisation",       10, "2026-01-05", "2026-01-15",  100),
        ("A1010", "Excavation & Earthworks", 25, "2026-01-16", "2026-02-10",   75),
        ("A1020", "Foundations",             30, "2026-02-11", "2026-03-13",   40),
        ("A1030", "Superstructure",          90, "2026-03-14", "2026-06-12",    0),
        ("A1040", "Roofing",                 20, "2026-06-13", "2026-07-03",    0),
    ]
    for r in rows:
        ws.append(list(r))
    return wb


# ---------------------------------------------------------------------------
# 1. Happy path
# ---------------------------------------------------------------------------

def test_parse_xlsx_schedule_happy_path():
    container = ConstructionContainer()
    path = _write_workbook(_basic_schedule_workbook())
    try:
        result = container._parse_xlsx_schedule(path)
    finally:
        os.unlink(path)

    assert result["status"] == "success"
    assert result["engine"] == "xlsx-heuristic"
    assert result["sheet"] == "Project Schedule"
    assert len(result["activities"]) == 5

    summary = result["summary"]
    assert summary["total_activities"] == 5
    assert summary["with_duration"] == 5
    assert summary["with_dates"] == 5
    assert summary["average_duration_days"] > 0

    a0 = result["activities"][0]
    # Each activity has the contract-shape fields.
    for k in ("task_id", "task_name", "duration", "target_start_date", "target_end_date"):
        assert k in a0, f"missing field {k}"
    assert a0["duration"] > 0
    # Dates parsed to ISO-8601 strings.
    assert isinstance(a0["target_start_date"], str)
    assert isinstance(a0["target_end_date"], str)
    # task_name preserved.
    names = {a["task_name"] for a in result["activities"]}
    assert "Site Mobilisation" in names


# ---------------------------------------------------------------------------
# 2. No schedule-shaped sheet
# ---------------------------------------------------------------------------

def test_parse_xlsx_schedule_no_schedule_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Random"
    ws.append(["Foo", "Bar", "Baz"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    path = _write_workbook(wb)

    container = ConstructionContainer()
    try:
        result = container._parse_xlsx_schedule(path)
    finally:
        os.unlink(path)

    assert result["status"] == "error"
    assert "no schedule-shaped sheet found" in result["error"]
    assert result["sheets_inspected"] == ["Random"]


# ---------------------------------------------------------------------------
# 3. Multi-sheet — picks the schedule-shaped one even when not first
# ---------------------------------------------------------------------------

def test_parse_xlsx_schedule_multi_sheet_picks_correct_sheet():
    wb = Workbook()
    inputs = wb.active
    inputs.title = "Inputs"
    inputs.append(["Parameter", "Value", "Notes"])
    inputs.append(["Currency", "USD", ""])
    inputs.append(["Region", "NA", ""])

    wbs_sheet = wb.create_sheet("WBS")
    wbs_sheet.append(["WBS", "Task Name", "Duration", "Start", "Finish"])
    wbs_sheet.append(["1.1", "Engineering", 30, "2026-02-01", "2026-03-03"])
    wbs_sheet.append(["1.2", "Procurement", 60, "2026-03-04", "2026-05-03"])
    wbs_sheet.append(["1.3", "Construction", 120, "2026-05-04", "2026-09-01"])

    path = _write_workbook(wb)

    container = ConstructionContainer()
    try:
        result = container._parse_xlsx_schedule(path)
    finally:
        os.unlink(path)

    assert result["status"] == "success"
    assert result["sheet"] == "WBS"
    assert len(result["activities"]) == 3
    # WBS column should propagate to task_code/task_id.
    ids = {a["task_id"] for a in result["activities"]}
    assert ids == {"1.1", "1.2", "1.3"}


# ---------------------------------------------------------------------------
# 4. Native datetime cell values serialise to ISO-8601
# ---------------------------------------------------------------------------

def test_parse_xlsx_schedule_datetime_cell_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["Activity ID", "Activity Name", "Duration", "Start", "Finish"])
    ws.append([
        "T1",
        "Kickoff",
        5,
        datetime(2026, 4, 1, 9, 0, 0),
        datetime(2026, 4, 6, 17, 0, 0),
    ])
    path = _write_workbook(wb)

    container = ConstructionContainer()
    try:
        result = container._parse_xlsx_schedule(path)
    finally:
        os.unlink(path)

    assert result["status"] == "success"
    assert len(result["activities"]) == 1
    a = result["activities"][0]
    assert isinstance(a["target_start_date"], str)
    assert isinstance(a["target_end_date"], str)
    assert a["target_start_date"].startswith("2026-04-01")
    assert a["target_end_date"].startswith("2026-04-06")


# ---------------------------------------------------------------------------
# 5. Dispatch from parse_primavera_schedule for .xlsx
# ---------------------------------------------------------------------------

def test_parse_primavera_schedule_dispatches_xlsx():
    path = _write_workbook(_basic_schedule_workbook())
    container = ConstructionContainer()
    try:
        result = asyncio.run(
            container.parse_primavera_schedule(
                {"file_path": path}, {"include_details": True}
            )
        )
    finally:
        os.unlink(path)

    assert result["status"] == "success"
    assert result["action"] == "schedule_analysis"
    # 5 activities round-trip into the public summary.
    assert result["summary"]["total_activities"] == 5
    assert result["detailed_activities"] is not None
    assert len(result["detailed_activities"]) == 5


# ---------------------------------------------------------------------------
# 6. Classify XLSX as schedule based on content
# ---------------------------------------------------------------------------

def test_classify_document_detects_xlsx_schedule_by_sheet_shape():
    """Even with a neutral filename, an XLSX whose sheet looks schedule-shaped
    must classify as `schedule` so process_document dispatches correctly."""
    wb = _basic_schedule_workbook()
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="anonymous_workbook_")
    os.close(fd)
    wb.save(path)

    container = ConstructionContainer()
    try:
        doc_type = asyncio.run(container._classify_document(path))
    finally:
        os.unlink(path)

    assert doc_type == "schedule"


def test_classify_document_detects_xlsx_schedule_by_filename():
    """Filename hint alone is enough — no need to crack the workbook open."""
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="L2_Schedule_Anthropic_")
    os.close(fd)
    # Empty workbook — header inspection would fail, filename hint wins.
    Workbook().save(path)

    container = ConstructionContainer()
    try:
        doc_type = asyncio.run(container._classify_document(path))
    finally:
        os.unlink(path)

    assert doc_type == "schedule"
